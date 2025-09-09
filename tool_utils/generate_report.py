import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import os
import traceback

# 读取JSON数据
def read_json_file(json_path):
    if not os.path.exists(json_path):
        data = None
    else:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    return data

# 将数据写入已有的Excel文件
def write_to_excel(data, excel_path):
    if(data == None):
        return
    df = pd.DataFrame(data)
    with pd.ExcelWriter(excel_path, mode='a', if_exists='replace') as writer:
        df.to_excel(writer, index=False)


# 初始化
def generate_report(json_path, excel_path, report_template):
    try:
        rlt_res = {
        "code_id": 1,
        "describe": "data clean and overview",
        "out": {},
        "summary": [],
        "table": [],
        "ErrorCode": [0, 0, '']
    }
        json_data = read_json_file(json_path)
        if json_data == None:
            return
        # print('这是文件中的json数据：', json_data)
        # print('这是读取到文件数据的数据类型：', type(json_data))
        data = json_data['data']
        json_str = json.dumps(data)
        data_arr = json.loads(json_str)

        # 加载已存在的Excel文件
        wb = load_workbook(report_template)

        # 选择要操作的工作表
        ws = wb.active

        # 写入数据到指定单元格
        ws['D5'] = str(data_arr['car_info']['车辆识别码']['结果值'])
        ws['D7'] = ILLEGAL_CHARACTERS_RE.sub(r'', str(data_arr['car_info']['生产厂商']['结果值']))
        manufacture_date = str(data_arr['car_info']['生产日期']['结果值'])
        ws['D8'] = manufacture_date if manufacture_date != 'NaN' else ''

        ws['D10'] = str(data_arr['battery_info']['电池类型']['结果值'])
        ws['D11'] = str(data_arr['battery_info']['电芯数量']['结果值'])
        ws['D12'] = str(data_arr['battery_info']['额定容量']['结果值'])
        ws['D13'] = str(data_arr['battery_info']['额定能量']['结果值'])

        alignment_type = Alignment(horizontal="left")  # 水平对齐方式。center居中；left左；right右

        ws['D15'].alignment = alignment_type
        if '充电SOC区间' in data_arr['charging_info']:
            ws['D15'] = str(data_arr['charging_info']['充电SOC区间']['结果值'])

        ws['D16'].alignment = alignment_type
        if '充电总电压区间' in data_arr['charging_info']:
            ws['D16'] = str(data_arr['charging_info']['充电总电压区间']['结果值'])

        ws['D17'].alignment = alignment_type
        if '单体电压区间' in data_arr['charging_info']:
            ws['D17'] = str(data_arr['charging_info']['单体电压区间']['结果值'])

        ws['D18'].alignment = alignment_type
        if '充电电流区间' in data_arr['charging_info']:
            ws['D18'] = str(data_arr['charging_info']['充电电流区间']['结果值'])

        ws['D19'].alignment = alignment_type
        if '温度区间' in data_arr['charging_info']:
            ws['D19'] = str(data_arr['charging_info']['温度区间']['结果值'])

        ws['D20'].alignment = alignment_type
        if '当次充电容量' in data_arr['charging_info']:
            ws['D20'] = str(data_arr['charging_info']['当次充电容量']['结果值'])

        ws['D21'].alignment = alignment_type
        if '当次充电度数' in data_arr['charging_info']:
            ws['D21'] = str(data_arr['charging_info']['当次充电度数']['结果值'])

        ws['D22'].alignment = alignment_type
        if '当次充电可行驶里程数' in data_arr['charging_info']:
            ws['D22'] = str(data_arr['charging_info']['当次充电可行驶里程数']['结果值'])

        ######

        if '蓄电池最高单体电芯剩余容量' in data_arr['soh_info']:
            ws['D25'] = str(data_arr['soh_info']['蓄电池最高单体电芯剩余容量']['结果值'])
            ws['E25'] = str(data_arr['soh_info']['蓄电池最高单体电芯剩余容量']['异常提示'])
            ws['F25'] = str(data_arr['soh_info']['蓄电池最高单体电芯剩余容量']['参考范围'])
            ws['G25'] = str(data_arr['soh_info']['蓄电池最高单体电芯剩余容量']['单位'])
            ws['H25'] = str(data_arr['soh_info']['蓄电池最高单体电芯剩余容量']['指标说明'])

        if '蓄电池剩余总能量' in data_arr['soh_info']:
            ws['D26'] = str(data_arr['soh_info']['蓄电池剩余总能量']['结果值'])
            ws['E26'] = str(data_arr['soh_info']['蓄电池剩余总能量']['异常提示'])
            ws['F26'] = str(data_arr['soh_info']['蓄电池剩余总能量']['参考范围'])
            ws['G26'] = str(data_arr['soh_info']['蓄电池剩余总能量']['单位'])
            ws['H26'] = str(data_arr['soh_info']['蓄电池剩余总能量']['指标说明'])

        if '蓄电池平均容量保有率' in data_arr['soh_info']:
            ws['D27'] = str(data_arr['soh_info']['蓄电池平均容量保有率']['结果值'])
            ws['E27'] = str(data_arr['soh_info']['蓄电池平均容量保有率']['异常提示'])
            ws['F27'] = str(data_arr['soh_info']['蓄电池平均容量保有率']['参考范围'])
            ws['G27'] = str(data_arr['soh_info']['蓄电池平均容量保有率']['单位'])
            ws['H27'] = str(data_arr['soh_info']['蓄电池平均容量保有率']['指标说明'])

        if '蓄电池容量损失速率' in data_arr['soh_info']:
            ws['D28'] = str(data_arr['soh_info']['蓄电池容量损失速率']['结果值'])
            ws['E28'] = str(data_arr['soh_info']['蓄电池容量损失速率']['异常提示'])
            ws['F28'] = str(data_arr['soh_info']['蓄电池容量损失速率']['参考范围'])
            ws['G28'] = str(data_arr['soh_info']['蓄电池容量损失速率']['单位'])
            ws['H28'] = str(data_arr['soh_info']['蓄电池容量损失速率']['指标说明'])

        ######

        if '蓄电池总压最大上升速率' in data_arr['voltage_info']:
            ws['D29'] = str(data_arr['voltage_info']['蓄电池总压最大上升速率']['结果值'])
            ws['E29'] = str(data_arr['voltage_info']['蓄电池总压最大上升速率']['异常提示'])
            ws['F29'] = str(data_arr['voltage_info']['蓄电池总压最大上升速率']['参考范围'])
            ws['G29'] = str(data_arr['voltage_info']['蓄电池总压最大上升速率']['单位'])
            ws['H29'] = str(data_arr['voltage_info']['蓄电池总压最大上升速率']['指标说明'])

        if '蓄电池单体最大压差值' in data_arr['voltage_info']:
            ws['D30'] = str(data_arr['voltage_info']['蓄电池单体最大压差值']['结果值'])
            ws['E30'] = str(data_arr['voltage_info']['蓄电池单体最大压差值']['异常提示'])
            ws['F30'] = str(data_arr['voltage_info']['蓄电池单体最大压差值']['参考范围'])
            ws['G30'] = str(data_arr['voltage_info']['蓄电池单体最大压差值']['单位'])
            ws['H30'] = str(data_arr['voltage_info']['蓄电池单体最大压差值']['指标说明'])

        if '蓄电池单体电压最大上升速率' in data_arr['voltage_info']:
            ws['D31'] = str(data_arr['voltage_info']['蓄电池单体电压最大上升速率']['结果值'])
            ws['E31'] = str(data_arr['voltage_info']['蓄电池单体电压最大上升速率']['异常提示'])
            ws['F31'] = str(data_arr['voltage_info']['蓄电池单体电压最大上升速率']['参考范围'])
            ws['G31'] = str(data_arr['voltage_info']['蓄电池单体电压最大上升速率']['单位'])
            ws['H31'] = str(data_arr['voltage_info']['蓄电池单体电压最大上升速率']['指标说明'])

        if '蓄电池组电压一致性' in data_arr['voltage_info']:
            ws['D32'] = str(data_arr['voltage_info']['蓄电池组电压一致性']['结果值'])
            ws['E32'] = str(data_arr['voltage_info']['蓄电池组电压一致性']['异常提示'])
            ws['F32'] = str(data_arr['voltage_info']['蓄电池组电压一致性']['参考范围'])
            ws['G32'] = str(data_arr['voltage_info']['蓄电池组电压一致性']['单位'])
            ws['H32'] = str(data_arr['voltage_info']['蓄电池组电压一致性']['指标说明'])

        if '蓄电池容压比一致性' in data_arr['voltage_info']:
            ws['D33'] = str(data_arr['voltage_info']['蓄电池容压比一致性']['结果值'])
            ws['E33'] = str(data_arr['voltage_info']['蓄电池容压比一致性']['异常提示'])
            ws['F33'] = str(data_arr['voltage_info']['蓄电池容压比一致性']['参考范围'])
            ws['G33'] = str(data_arr['voltage_info']['蓄电池容压比一致性']['单位'])
            ws['H33'] = str(data_arr['voltage_info']['蓄电池容压比一致性']['指标说明'])

        ######

        if '最大电流上升速率' in data_arr['current_info']:
            ws['D34'] = str(data_arr['current_info']['最大电流上升速率']['结果值'])
            ws['E34'] = str(data_arr['current_info']['最大电流上升速率']['异常提示'])
            ws['F34'] = str(data_arr['current_info']['最大电流上升速率']['参考范围'])
            ws['G34'] = str(data_arr['current_info']['最大电流上升速率']['单位'])
            ws['H34'] = str(data_arr['current_info']['最大电流上升速率']['指标说明'])

        if '快充时长' in data_arr['current_info']:
            ws['D35'] = str(data_arr['current_info']['快充时长']['结果值'])
            ws['E35'] = str(data_arr['current_info']['快充时长']['异常提示'])
            ws['F35'] = str(data_arr['current_info']['快充时长']['参考范围'])
            ws['G35'] = str(data_arr['current_info']['快充时长']['单位'])
            ws['H35'] = str(data_arr['current_info']['快充时长']['指标说明'])

        ######

        if '温度探头检测最大差值' in data_arr['temperature_info']:
            ws['D36'] = str(data_arr['temperature_info']['温度探头检测最大差值']['结果值'])
            ws['E36'] = str(data_arr['temperature_info']['温度探头检测最大差值']['异常提示'])
            ws['F36'] = str(data_arr['temperature_info']['温度探头检测最大差值']['参考范围'])
            ws['G36'] = str(data_arr['temperature_info']['温度探头检测最大差值']['单位'])
            ws['H36'] = str(data_arr['temperature_info']['温度探头检测最大差值']['指标说明'])

        if '温度探头最大上升速率' in data_arr['temperature_info']:
            ws['D37'] = str(data_arr['temperature_info']['温度探头最大上升速率']['结果值'])
            ws['E37'] = str(data_arr['temperature_info']['温度探头最大上升速率']['异常提示'])
            ws['F37'] = str(data_arr['temperature_info']['温度探头最大上升速率']['参考范围'])
            ws['G37'] = str(data_arr['temperature_info']['温度探头最大上升速率']['单位'])
            ws['H37'] = str(data_arr['temperature_info']['温度探头最大上升速率']['指标说明'])

        if '蓄电池组温度一致性' in data_arr['temperature_info']:
            ws['D38'] = str(data_arr['temperature_info']['蓄电池组温度一致性']['结果值'])
            ws['E38'] = str(data_arr['temperature_info']['蓄电池组温度一致性']['异常提示'])
            ws['F38'] = str(data_arr['temperature_info']['蓄电池组温度一致性']['参考范围'])
            ws['G38'] = str(data_arr['temperature_info']['蓄电池组温度一致性']['单位'])
            ws['H38'] = str(data_arr['temperature_info']['蓄电池组温度一致性']['指标说明'])

        ######

        if '电压值异常比例' in data_arr['abusive_info']:
            ws['D39'] = str(data_arr['abusive_info']['电压值异常比例']['结果值'])
            ws['E39'] = str(data_arr['abusive_info']['电压值异常比例']['异常提示'])
            ws['F39'] = str(data_arr['abusive_info']['电压值异常比例']['参考范围'])
            ws['G39'] = str(data_arr['abusive_info']['电压值异常比例']['单位'])
            ws['H39'] = str(data_arr['abusive_info']['电压值异常比例']['指标说明'])

        if '电流异常值比例' in data_arr['abusive_info']:
            ws['D40'] = str(data_arr['abusive_info']['电流异常值比例']['结果值'])
            ws['E40'] = str(data_arr['abusive_info']['电流异常值比例']['异常提示'])
            ws['F40'] = str(data_arr['abusive_info']['电流异常值比例']['参考范围'])
            ws['G40'] = str(data_arr['abusive_info']['电流异常值比例']['单位'])
            ws['H40'] = str(data_arr['abusive_info']['电流异常值比例']['指标说明'])

        if '温度值异常比例' in data_arr['abusive_info']:
            ws['D41'] = str(data_arr['abusive_info']['温度值异常比例']['结果值'])
            ws['E41'] = str(data_arr['abusive_info']['温度值异常比例']['异常提示'])
            ws['F41'] = str(data_arr['abusive_info']['温度值异常比例']['参考范围'])
            ws['G41'] = str(data_arr['abusive_info']['温度值异常比例']['单位'])
            ws['H41'] = str(data_arr['abusive_info']['温度值异常比例']['指标说明'])

        ######

        if '电压传感器有效性' in data_arr['sensor_info']:
            ws['D44'] = str(data_arr['sensor_info']['电压传感器有效性']['结果值'])
            ws['E44'] = str(data_arr['sensor_info']['电压传感器有效性']['异常提示'])
            ws['F44'] = str(data_arr['sensor_info']['电压传感器有效性']['参考范围'])
            ws['G44'] = str(data_arr['sensor_info']['电压传感器有效性']['单位'])
            ws['H44'] = str(data_arr['sensor_info']['电压传感器有效性']['指标说明'])

        if '温度传感器有效性' in data_arr['sensor_info']:
            ws['D45'] = str(data_arr['sensor_info']['温度传感器有效性']['结果值'])
            ws['E45'] = str(data_arr['sensor_info']['温度传感器有效性']['异常提示'])
            ws['F45'] = str(data_arr['sensor_info']['温度传感器有效性']['参考范围'])
            ws['G45'] = str(data_arr['sensor_info']['温度传感器有效性']['单位'])
            ws['H45'] = str(data_arr['sensor_info']['温度传感器有效性']['指标说明'])

        if '电流传感器有效性' in data_arr['sensor_info']:
            ws['D46'] = str(data_arr['sensor_info']['电流传感器有效性']['结果值'])
            ws['E46'] = str(data_arr['sensor_info']['电流传感器有效性']['异常提示'])
            ws['F46'] = str(data_arr['sensor_info']['电流传感器有效性']['参考范围'])
            ws['G46'] = str(data_arr['sensor_info']['电流传感器有效性']['单位'])
            ws['H46'] = str(data_arr['sensor_info']['电流传感器有效性']['指标说明'])

        ######

        if '单体电池电压报警' in data_arr['alarm_management_info']:
            ws['D47'] = str(data_arr['alarm_management_info']['单体电池电压报警']['结果值'])
            ws['E47'] = str(data_arr['alarm_management_info']['单体电池电压报警']['异常提示'])
            ws['F47'] = str(data_arr['alarm_management_info']['单体电池电压报警']['参考范围'])
            ws['G47'] = str(data_arr['alarm_management_info']['单体电池电压报警']['单位'])
            ws['H47'] = str(data_arr['alarm_management_info']['单体电池电压报警']['指标说明'])

        if '充电过温' in data_arr['alarm_management_info']:
            ws['D48'] = str(data_arr['alarm_management_info']['充电过温']['结果值'])
            ws['E48'] = str(data_arr['alarm_management_info']['充电过温']['异常提示'])
            ws['F48'] = str(data_arr['alarm_management_info']['充电过温']['参考范围'])
            ws['G48'] = str(data_arr['alarm_management_info']['充电过温']['单位'])
            ws['H48'] = str(data_arr['alarm_management_info']['充电过温']['指标说明'])

        if '充电过电流' in data_arr['alarm_management_info']:
            ws['D49'] = str(data_arr['alarm_management_info']['充电过电流']['结果值'])
            ws['E49'] = str(data_arr['alarm_management_info']['充电过电流']['异常提示'])
            ws['F49'] = str(data_arr['alarm_management_info']['充电过电流']['参考范围'])
            ws['G49'] = str(data_arr['alarm_management_info']['充电过电流']['单位'])
            ws['H49'] = str(data_arr['alarm_management_info']['充电过电流']['指标说明'])

        if '电池绝缘状态' in data_arr['alarm_management_info']:
            ws['D50'] = str(data_arr['alarm_management_info']['电池绝缘状态']['结果值'])
            ws['E50'] = str(data_arr['alarm_management_info']['电池绝缘状态']['异常提示'])
            ws['F50'] = str(data_arr['alarm_management_info']['电池绝缘状态']['参考范围'])
            ws['G50'] = str(data_arr['alarm_management_info']['电池绝缘状态']['单位'])
            ws['H50'] = str(data_arr['alarm_management_info']['电池绝缘状态']['指标说明'])


        if '输出连接器连接状态' in data_arr['alarm_management_info']:
            ws['D51'] = str(data_arr['alarm_management_info']['输出连接器连接状态']['结果值'])
            ws['E51'] = str(data_arr['alarm_management_info']['输出连接器连接状态']['异常提示'])
            ws['F51'] = str(data_arr['alarm_management_info']['输出连接器连接状态']['参考范围'])
            ws['G51'] = str(data_arr['alarm_management_info']['输出连接器连接状态']['单位'])
            ws['H51'] = str(data_arr['alarm_management_info']['输出连接器连接状态']['指标说明'])

        if 'BMS中止充电原因' in data_arr['alarm_management_info']:
            ws['D52'] = str(data_arr['alarm_management_info']['BMS中止充电原因']['结果值'])
            ws['E52'] = str(data_arr['alarm_management_info']['BMS中止充电原因']['异常提示'])
            ws['F52'] = str(data_arr['alarm_management_info']['BMS中止充电原因']['参考范围'])
            ws['G52'] = str(data_arr['alarm_management_info']['BMS中止充电原因']['单位'])
            ws['H52'] = str(data_arr['alarm_management_info']['BMS中止充电原因']['指标说明'])

        if '充电机中止充电原因' in data_arr['alarm_management_info']:
            ws['D53'] = str(data_arr['alarm_management_info']['充电机中止充电原因']['结果值'])
            ws['E53'] = str(data_arr['alarm_management_info']['充电机中止充电原因']['异常提示'])
            ws['F53'] = str(data_arr['alarm_management_info']['充电机中止充电原因']['参考范围'])
            ws['G53'] = str(data_arr['alarm_management_info']['充电机中止充电原因']['单位'])
            ws['H53'] = str(data_arr['alarm_management_info']['充电机中止充电原因']['指标说明'])

        if '电池容量保有率评分' in data_arr['score_info']:
            ws['C56'] = str(data_arr['score_info']['电池容量保有率评分']['结果值'])

        if '充电工况评分' in data_arr['score_info']:
            ws['C57'] = str(data_arr['score_info']['充电工况评分']['结果值'])

        if '电池组性能一致性评分' in data_arr['score_info']:
            ws['C58'] = str(data_arr['score_info']['电池组性能一致性评分']['结果值'])

        if '车载BMS管理系统评分' in data_arr['score_info']:
            ws['C59'] = str(data_arr['score_info']['车载BMS管理系统评分']['结果值'])

        if '充电安全评分' in data_arr['score_info']:
            ws['C60'] = str(data_arr['score_info']['充电安全评分']['结果值'])

        if '总评分' in data_arr['score_info']:
            ws['C61'] = str(data_arr['score_info']['总评分']['结果值'])

        # 保存修改后的工作簿
        wb.save(excel_path)
        return  rlt_res
    except Exception as e:
        rlt_res['ErrorCode'][0] = -99
        rlt_res['ErrorCode'][2] = e
        traceback.print_exc()
        return rlt_res

# if __name__ == "__main__":
#     current_cwd = os.getcwd()    
#     json_path = os.path.join(current_cwd, 'report', '20241212', '0000132082600010001012024121305485957', 'data.json')
#     excel_path = os.path.join(current_cwd, 'report', '20241212', '0000132082600010001012024121305485957', 'report.xlsx')
#     report_template = os.path.join(current_cwd, 'report', 'report.xlsx')

#     print(json_path)
#     print(excel_path)
#     print(report_template)
#     init(json_path, excel_path, report_template)

